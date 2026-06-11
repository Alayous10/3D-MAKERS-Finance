# -*- coding: utf-8 -*-
"""
3D Makers Finance — Professional Excel Generator
Generates 3DMAkers_Finance.xlsx with 7 sheets, real Excel formulas,
data validation, professional formatting, and sample data.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

# ═══════════════════════════════════════════════════════════
# CONSTANTS & STYLE DEFINITIONS
# ═══════════════════════════════════════════════════════════

FILENAME = "3DMAkers_Finance.xlsx"

# Colors
DARK_HEADER_BG = "1B2A4A"       # Deep navy
WHITE = "FFFFFF"
YELLOW_BG = "FFF9C4"            # Light yellow for editable assumptions
GRAY_TOTAL_BG = "E8E8E8"        # Light gray for total rows
BLUE_MANUAL = "1565C0"          # Blue for manual input
GREEN_CROSS_REF = "2E7D32"      # Green for cross-sheet references
BLACK_AUTO = "000000"           # Black for auto-calculated

# Fonts
FONT_HEADER = Font(name="Arial", bold=True, color=WHITE, size=11)
FONT_TITLE = Font(name="Arial", bold=True, color=WHITE, size=14)
FONT_SECTION = Font(name="Arial", bold=True, color=DARK_HEADER_BG, size=12)
FONT_NORMAL = Font(name="Arial", color=BLACK_AUTO, size=10)
FONT_MANUAL = Font(name="Arial", color=BLUE_MANUAL, size=10)
FONT_FORMULA = Font(name="Arial", color=BLACK_AUTO, size=10)
FONT_CROSS_REF = Font(name="Arial", color=GREEN_CROSS_REF, size=10, bold=True)
FONT_TOTAL = Font(name="Arial", bold=True, color=BLACK_AUTO, size=11)
FONT_KPI_LABEL = Font(name="Arial", color="37474F", size=11)
FONT_KPI_VALUE = Font(name="Arial", bold=True, color=DARK_HEADER_BG, size=13)
FONT_KPI_VALUE_GREEN = Font(name="Arial", bold=True, color=GREEN_CROSS_REF, size=13)
FONT_ASSUMPTION_LABEL = Font(name="Arial", bold=True, color="5D4037", size=10)
FONT_ASSUMPTION_VAL = Font(name="Arial", bold=True, color=BLUE_MANUAL, size=11)
FONT_NOTE = Font(name="Arial", italic=True, color="757575", size=9)

# Fills
FILL_HEADER = PatternFill(start_color=DARK_HEADER_BG, end_color=DARK_HEADER_BG, fill_type="solid")
FILL_YELLOW = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid")
FILL_TOTAL = PatternFill(start_color=GRAY_TOTAL_BG, end_color=GRAY_TOTAL_BG, fill_type="solid")
FILL_DASHBOARD_BG = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
FILL_SECTION_BG = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

# Alignment
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Borders
THIN_BORDER = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color=DARK_HEADER_BG))

# Number formats
FMT_CURRENCY = '$#,##0.00_);($#,##0.00)'
FMT_PERCENT = '0.0%'
FMT_NUMBER = '#,##0'
FMT_DATE = 'YYYY-MM-DD'
FMT_DECIMAL = '#,##0.00'


# ═══════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════

def set_col_widths(ws, widths_dict):
    """Set column widths. widths_dict = {col_letter: width}"""
    for col, w in widths_dict.items():
        ws.column_dimensions[col].width = w


def style_header_row(ws, row, max_col):
    """Apply dark header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def style_total_row(ws, row, max_col):
    """Apply total row styling."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_TOTAL
        cell.fill = FILL_TOTAL
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, is_manual=False, is_formula=False, is_cross_ref=False):
    """Apply appropriate font color to a data cell."""
    cell = ws.cell(row=row, column=col)
    if is_cross_ref:
        cell.font = FONT_CROSS_REF
    elif is_manual:
        cell.font = FONT_MANUAL
    elif is_formula:
        cell.font = FONT_FORMULA
    else:
        cell.font = FONT_NORMAL
    cell.border = THIN_BORDER
    cell.alignment = ALIGN_CENTER
    return cell


def add_data_rows_formatting(ws, start_row, end_row, max_col, manual_cols=None, formula_cols=None):
    """Format a range of data rows."""
    if manual_cols is None:
        manual_cols = []
    if formula_cols is None:
        formula_cols = []
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            is_manual = c in manual_cols
            is_formula = c in formula_cols
            style_data_cell(ws, r, c, is_manual=is_manual, is_formula=is_formula)


# ═══════════════════════════════════════════════════════════
# SHEET BUILDERS
# ═══════════════════════════════════════════════════════════

def build_capital(wb):
    """Sheet 2: رأس المال (Capital)"""
    ws = wb.create_sheet("Capital")

    # Title
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "رأس المال — Capital Investments"
    title_cell.font = FONT_TITLE
    title_cell.fill = FILL_HEADER
    title_cell.alignment = ALIGN_CENTER

    # Headers - Row 2
    headers = ["التاريخ\nDate", "نوع الاستثمار\nInvestment Type", "الوصف\nDescription",
               "المبلغ ($)\nAmount", "ملاحظات\nNotes"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, 5)

    # Data Validation — Investment Type dropdown
    inv_types = '"طابعة ثلاثية الأبعاد,موقع إلكتروني / منصة,معدات وأدوات,ترخيص وتصاريح,أخرى"'
    dv = DataValidation(type="list", formula1=inv_types, allow_blank=True)
    dv.error = "الرجاء اختيار نوع من القائمة"
    dv.errorTitle = "خطأ في الإدخال"
    dv.prompt = "اختر نوع الاستثمار"
    dv.promptTitle = "نوع الاستثمار"
    ws.add_data_validation(dv)
    dv.add("B3:B102")

    # Sample data — Row 3
    ws.cell(row=3, column=1, value=date(2025, 1, 15))
    ws.cell(row=3, column=2, value="طابعة ثلاثية الأبعاد")
    ws.cell(row=3, column=3, value="Creality Ender 3 V3 SE")
    ws.cell(row=3, column=4, value=250)
    ws.cell(row=3, column=5, value="الدفعة الأولى — ثمن الطابعة كاملاً")

    ws.cell(row=4, column=1, value=date(2025, 2, 1))
    ws.cell(row=4, column=2, value="موقع إلكتروني / منصة")
    ws.cell(row=4, column=3, value="بناء واستضافة المتجر الإلكتروني")
    ws.cell(row=4, column=4, value=150)
    ws.cell(row=4, column=5, value="تصميم + استضافة سنة")

    ws.cell(row=5, column=1, value=date(2025, 2, 10))
    ws.cell(row=5, column=2, value="معدات وأدوات")
    ws.cell(row=5, column=3, value="أدوات تشطيب + مواد تغليف")
    ws.cell(row=5, column=4, value=35)
    ws.cell(row=5, column=5, value="مقص حراري + ورق صنفرة + أكياس")

    # Format data rows
    for r in range(3, 103):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.font = FONT_MANUAL
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
            if c == 1:
                cell.number_format = FMT_DATE
            elif c == 4:
                cell.number_format = FMT_CURRENCY

    # Total row — Row 103
    total_row = 103
    ws.cell(row=total_row, column=1, value="الإجمالي / Total")
    ws.cell(row=total_row, column=4, value="=SUM(D3:D102)")
    ws.cell(row=total_row, column=4).number_format = FMT_CURRENCY
    style_total_row(ws, total_row, 5)

    # Column widths
    set_col_widths(ws, {"A": 15, "B": 25, "C": 30, "D": 18, "E": 30})

    # Freeze panes
    ws.freeze_panes = "A3"

    return ws


def build_product_calculator(wb):
    """Sheet 3: حاسبة التكلفة والسعر (Product Cost Calculator)"""
    ws = wb.create_sheet("ProductCalc")

    # Title row
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "حاسبة التكلفة والسعر — Product Cost Calculator"
    title_cell.font = FONT_TITLE
    title_cell.fill = FILL_HEADER
    title_cell.alignment = ALIGN_CENTER

    # Assumptions section — Row 2-3 (yellow background)
    ws.cell(row=2, column=1, value="الافتراضات / Assumptions").font = FONT_ASSUMPTION_LABEL
    ws.merge_cells("A2:B2")

    ws.cell(row=3, column=1, value="سعر الفيلمنت لكل غرام ($)")
    ws.cell(row=3, column=1).font = FONT_ASSUMPTION_LABEL
    cell_filament = ws.cell(row=3, column=2, value=0.015)
    cell_filament.font = FONT_ASSUMPTION_VAL
    cell_filament.fill = FILL_YELLOW
    cell_filament.number_format = '$#,##0.000'
    cell_filament.border = THIN_BORDER

    ws.cell(row=4, column=1, value="تكلفة الطابعة لكل ساعة ($)")
    ws.cell(row=4, column=1).font = FONT_ASSUMPTION_LABEL
    cell_printer = ws.cell(row=4, column=2, value=0.30)
    cell_printer.font = FONT_ASSUMPTION_VAL
    cell_printer.fill = FILL_YELLOW
    cell_printer.number_format = FMT_CURRENCY
    cell_printer.border = THIN_BORDER

    # Note
    ws.cell(row=3, column=4, value="🟨 الخلايا الصفراء = قيم قابلة للتعديل").font = FONT_NOTE
    ws.cell(row=4, column=4, value="جميع الصيغ مرتبطة بهذه الخانتين تلقائياً").font = FONT_NOTE

    # Headers — Row 6
    headers = [
        "اسم المنتج\nProduct Name",           # A
        "وزن الفيلمنت (غرام)\nFilament (g)",   # B
        "تكلفة الفيلمنت ($)\nFilament Cost",   # C
        "وقت الطباعة (ساعة)\nPrint Time (h)",  # D
        "تكلفة التشغيل ($)\nPrinter Cost",     # E
        "مواد إضافية ($)\nExtra Materials",     # F
        "إجمالي التكلفة ($)\nTotal Cost",       # G
        "هامش الربح المرغوب %\nDesired Margin", # H
        "سعر البيع المقترح ($)\nSuggested Price", # I
        "سعر البيع الفعلي ($)\nActual Price",   # J
        "الربح الفعلي ($)\nActual Profit",      # K
        "هامش الربح الفعلي %\nActual Margin",   # L
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=6, column=i, value=h)
    style_header_row(ws, 6, 12)

    # Formulas for rows 7-56 (50 product slots)
    for r in range(7, 57):
        # C: Filament cost = B * $B$3
        ws.cell(row=r, column=3, value=f"=IF(B{r}=\"\",\"\",B{r}*$B$3)")
        ws.cell(row=r, column=3).number_format = FMT_CURRENCY
        ws.cell(row=r, column=3).font = FONT_FORMULA

        # E: Printer cost = D * $B$4
        ws.cell(row=r, column=5, value=f"=IF(D{r}=\"\",\"\",D{r}*$B$4)")
        ws.cell(row=r, column=5).number_format = FMT_CURRENCY
        ws.cell(row=r, column=5).font = FONT_FORMULA

        # G: Total cost = C + E + F
        ws.cell(row=r, column=7, value=f"=IF(B{r}=\"\",\"\",C{r}+E{r}+IF(F{r}=\"\",0,F{r}))")
        ws.cell(row=r, column=7).number_format = FMT_CURRENCY
        ws.cell(row=r, column=7).font = FONT_FORMULA

        # I: Suggested price = G * (1 + H)
        ws.cell(row=r, column=9, value=f"=IF(G{r}=\"\",\"\",G{r}*(1+IF(H{r}=\"\",0,H{r})))")
        ws.cell(row=r, column=9).number_format = FMT_CURRENCY
        ws.cell(row=r, column=9).font = FONT_FORMULA

        # K: Actual profit = J - G
        ws.cell(row=r, column=11, value=f"=IF(J{r}=\"\",\"\",J{r}-G{r})")
        ws.cell(row=r, column=11).number_format = FMT_CURRENCY
        ws.cell(row=r, column=11).font = FONT_FORMULA

        # L: Actual margin = K / J
        ws.cell(row=r, column=12, value=f"=IFERROR(K{r}/J{r},\"\")")
        ws.cell(row=r, column=12).number_format = FMT_PERCENT
        ws.cell(row=r, column=12).font = FONT_FORMULA

        # Style all cells
        for c in range(1, 13):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
            # Manual input columns: A(1), B(2), D(4), F(6), H(8), J(10)
            if c in [1, 2, 4, 6, 8, 10]:
                cell.font = FONT_MANUAL
            if c == 8:
                cell.number_format = FMT_PERCENT
            if c in [2]:
                cell.number_format = FMT_NUMBER
            if c in [4]:
                cell.number_format = FMT_DECIMAL
            if c in [6, 10]:
                cell.number_format = FMT_CURRENCY

    # Sample data — 3 products
    samples = [
        ("علاقة مفاتيح — Keychain", 25, 1.5, 0.20, 0.5, 2.50),
        ("مجسم ديكور صغير — Figurine", 80, 4.0, 0.50, 0.6, 8.00),
        ("حامل موبايل — Phone Stand", 120, 5.0, 0.50, 0.4, 10.00),
    ]
    for idx, (name, weight, hours, extra, margin, price) in enumerate(samples):
        r = 7 + idx
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=weight)
        ws.cell(row=r, column=4, value=hours)
        ws.cell(row=r, column=6, value=extra)
        ws.cell(row=r, column=8, value=margin)
        ws.cell(row=r, column=10, value=price)

    # Column widths
    set_col_widths(ws, {
        "A": 28, "B": 18, "C": 16, "D": 18, "E": 16,
        "F": 16, "G": 16, "H": 18, "I": 18, "J": 18, "K": 16, "L": 16
    })

    ws.freeze_panes = "A7"
    return ws


def build_purchases(wb):
    """Sheet 4: سجل المشتريات (Purchases)"""
    ws = wb.create_sheet("Purchases")

    # Title
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "سجل المشتريات — Purchases Log"
    title_cell.font = FONT_TITLE
    title_cell.fill = FILL_HEADER
    title_cell.alignment = ALIGN_CENTER

    # Headers
    headers = [
        "التاريخ\nDate", "نوع المادة\nMaterial", "الكمية (غرام/وحدة)\nQuantity",
        "السعر الإجمالي ($)\nTotal Price", "سعر الوحدة ($)\nUnit Price",
        "المورّد\nSupplier", "ملاحظات\nNotes"
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, 7)

    # Formulas for unit price — rows 3-202
    for r in range(3, 203):
        # E: Unit Price = D / C
        ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/C{r},\"\")")
        ws.cell(row=r, column=5).number_format = '$#,##0.000'
        ws.cell(row=r, column=5).font = FONT_FORMULA

        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
            if c != 5:
                cell.font = FONT_MANUAL
            if c == 1:
                cell.number_format = FMT_DATE
            if c == 4:
                cell.number_format = FMT_CURRENCY
            if c == 3:
                cell.number_format = FMT_NUMBER

    # Sample data
    ws.cell(row=3, column=1, value=date(2025, 2, 15))
    ws.cell(row=3, column=2, value="فيلمنت PLA — أبيض")
    ws.cell(row=3, column=3, value=1000)
    ws.cell(row=3, column=4, value=15)
    ws.cell(row=3, column=6, value="Amazon")
    ws.cell(row=3, column=7, value="أول كمية — 1 كيلو")

    ws.cell(row=4, column=1, value=date(2025, 3, 1))
    ws.cell(row=4, column=2, value="فيلمنت PLA — أسود")
    ws.cell(row=4, column=3, value=1000)
    ws.cell(row=4, column=4, value=15)
    ws.cell(row=4, column=6, value="Amazon")
    ws.cell(row=4, column=7, value="كيلو ثاني")

    # Total row
    total_row = 203
    ws.cell(row=total_row, column=1, value="إجمالي مصاريف المواد / Total")
    ws.cell(row=total_row, column=4, value="=SUM(D3:D202)")
    ws.cell(row=total_row, column=4).number_format = FMT_CURRENCY
    style_total_row(ws, total_row, 7)

    set_col_widths(ws, {"A": 15, "B": 22, "C": 18, "D": 18, "E": 16, "F": 18, "G": 25})
    ws.freeze_panes = "A3"
    return ws


def build_sales(wb):
    """Sheet 5: سجل المبيعات (Sales)"""
    ws = wb.create_sheet("Sales")

    # Title
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "سجل المبيعات — Sales Log"
    title_cell.font = FONT_TITLE
    title_cell.fill = FILL_HEADER
    title_cell.alignment = ALIGN_CENTER

    # Headers
    headers = [
        "التاريخ\nDate", "اسم المنتج\nProduct", "الكمية\nQuantity",
        "سعر البيع للوحدة ($)\nUnit Price", "الإجمالي ($)\nTotal",
        "قناة البيع\nSales Channel", "ملاحظات\nNotes"
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, 7)

    # Sales Channel dropdown
    channels = '"موقع إلكتروني,يد بيد,واتساب,أخرى"'
    dv_channel = DataValidation(type="list", formula1=channels, allow_blank=True)
    dv_channel.error = "الرجاء اختيار قناة البيع"
    dv_channel.prompt = "اختر قناة البيع"
    ws.add_data_validation(dv_channel)
    dv_channel.add("F3:F202")

    # Formulas — rows 3-202
    for r in range(3, 203):
        # E: Total = C * D
        ws.cell(row=r, column=5, value=f"=IF(C{r}=\"\",\"\",C{r}*D{r})")
        ws.cell(row=r, column=5).number_format = FMT_CURRENCY
        ws.cell(row=r, column=5).font = FONT_FORMULA

        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
            if c != 5:
                cell.font = FONT_MANUAL
            if c == 1:
                cell.number_format = FMT_DATE
            if c == 4:
                cell.number_format = FMT_CURRENCY
            if c == 3:
                cell.number_format = FMT_NUMBER

    # Sample data
    ws.cell(row=3, column=1, value=date(2025, 3, 5))
    ws.cell(row=3, column=2, value="علاقة مفاتيح — Keychain")
    ws.cell(row=3, column=3, value=3)
    ws.cell(row=3, column=4, value=2.50)
    ws.cell(row=3, column=6, value="واتساب")
    ws.cell(row=3, column=7, value="أول طلبية!")

    ws.cell(row=4, column=1, value=date(2025, 3, 10))
    ws.cell(row=4, column=2, value="حامل موبايل — Phone Stand")
    ws.cell(row=4, column=3, value=1)
    ws.cell(row=4, column=4, value=10.00)
    ws.cell(row=4, column=6, value="يد بيد")
    ws.cell(row=4, column=7, value="بيع مباشر لصديق")

    # Totals — Row 203-204
    total_row = 203
    ws.cell(row=total_row, column=1, value="إجمالي الإيرادات / Total Revenue")
    ws.cell(row=total_row, column=3, value="=SUM(C3:C202)")
    ws.cell(row=total_row, column=3).number_format = FMT_NUMBER
    ws.cell(row=total_row, column=5, value="=SUM(E3:E202)")
    ws.cell(row=total_row, column=5).number_format = FMT_CURRENCY
    style_total_row(ws, total_row, 7)

    ws.cell(row=204, column=1, value="إجمالي عدد الوحدات المباعة / Total Units")
    ws.cell(row=204, column=3, value="=SUM(C3:C202)")
    ws.cell(row=204, column=3).number_format = FMT_NUMBER
    style_total_row(ws, 204, 7)

    set_col_widths(ws, {"A": 15, "B": 28, "C": 14, "D": 20, "E": 16, "F": 18, "G": 25})
    ws.freeze_panes = "A3"
    return ws


def build_expenses(wb):
    """Sheet 6: المصاريف العامة (Expenses)"""
    ws = wb.create_sheet("Expenses")

    # Title
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "المصاريف العامة — General Expenses"
    title_cell.font = FONT_TITLE
    title_cell.fill = FILL_HEADER
    title_cell.alignment = ALIGN_CENTER

    # Headers
    headers = [
        "التاريخ\nDate", "الفئة\nCategory", "الوصف\nDescription",
        "المبلغ ($)\nAmount", "ملاحظات\nNotes"
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, 5)

    # Category dropdown
    categories = '"اشتراكات وخدمات رقمية,نقل وتوصيل,صيانة وإصلاح,مستلزمات مكتبية,أخرى"'
    dv_cat = DataValidation(type="list", formula1=categories, allow_blank=True)
    dv_cat.error = "الرجاء اختيار فئة من القائمة"
    dv_cat.prompt = "اختر فئة المصروف"
    ws.add_data_validation(dv_cat)
    dv_cat.add("B3:B202")

    # Format rows
    for r in range(3, 203):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.font = FONT_MANUAL
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
            if c == 1:
                cell.number_format = FMT_DATE
            if c == 4:
                cell.number_format = FMT_CURRENCY

    # Sample data
    ws.cell(row=3, column=1, value=date(2025, 3, 1))
    ws.cell(row=3, column=2, value="اشتراكات وخدمات رقمية")
    ws.cell(row=3, column=3, value="اشتراك Canva Pro شهري")
    ws.cell(row=3, column=4, value=12.99)
    ws.cell(row=3, column=5, value="لتصميم منشورات السوشال ميديا")

    ws.cell(row=4, column=1, value=date(2025, 3, 15))
    ws.cell(row=4, column=2, value="نقل وتوصيل")
    ws.cell(row=4, column=3, value="توصيل طلبية للعميل")
    ws.cell(row=4, column=4, value=5.00)
    ws.cell(row=4, column=5, value="")

    # Total row
    total_row = 203
    ws.cell(row=total_row, column=1, value="إجمالي المصاريف / Total")
    ws.cell(row=total_row, column=4, value="=SUM(D3:D202)")
    ws.cell(row=total_row, column=4).number_format = FMT_CURRENCY
    style_total_row(ws, total_row, 5)

    set_col_widths(ws, {"A": 15, "B": 25, "C": 30, "D": 18, "E": 30})
    ws.freeze_panes = "A3"
    return ws


def build_advertising(wb):
    """Sheet 7: الإعلانات (Advertising)"""
    ws = wb.create_sheet("Advertising")

    # Title
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "سجل الإعلانات — Advertising Log"
    title_cell.font = FONT_TITLE
    title_cell.fill = FILL_HEADER
    title_cell.alignment = ALIGN_CENTER

    # Headers
    headers = [
        "التاريخ\nDate", "المنصة\nPlatform", "نوع الحملة\nCampaign Type",
        "الميزانية المخصصة ($)\nBudget", "المبلغ الفعلي ($)\nActual Spend",
        "عدد المبيعات الناتجة\nSales Count", "الإيراد الناتج ($)\nRevenue",
        "العائد على الإعلان\nROAS", "ملاحظات\nNotes"
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, 9)

    # Formulas — rows 3-102
    for r in range(3, 103):
        # H: ROAS = G / E
        ws.cell(row=r, column=8, value=f"=IFERROR(G{r}/E{r},\"\")")
        ws.cell(row=r, column=8).number_format = FMT_DECIMAL
        ws.cell(row=r, column=8).font = FONT_FORMULA

        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
            if c != 8:
                cell.font = FONT_MANUAL
            if c == 1:
                cell.number_format = FMT_DATE
            if c in [4, 5, 7]:
                cell.number_format = FMT_CURRENCY
            if c == 6:
                cell.number_format = FMT_NUMBER

    # Sample data
    ws.cell(row=3, column=1, value=date(2025, 4, 1))
    ws.cell(row=3, column=2, value="Instagram")
    ws.cell(row=3, column=3, value="ترويج منشور — Boost Post")
    ws.cell(row=3, column=4, value=20)
    ws.cell(row=3, column=5, value=18.50)
    ws.cell(row=3, column=6, value=4)
    ws.cell(row=3, column=7, value=35)
    ws.cell(row=3, column=9, value="أول حملة تجريبية")

    # Totals — Row 103-104
    total_row = 103
    ws.cell(row=total_row, column=1, value="إجمالي مصاريف الإعلانات / Total Ad Spend")
    ws.cell(row=total_row, column=5, value="=SUM(E3:E102)")
    ws.cell(row=total_row, column=5).number_format = FMT_CURRENCY
    ws.cell(row=total_row, column=7, value="=SUM(G3:G102)")
    ws.cell(row=total_row, column=7).number_format = FMT_CURRENCY
    style_total_row(ws, total_row, 9)

    ws.cell(row=104, column=1, value="متوسط ROAS / Average ROAS")
    ws.cell(row=104, column=8, value="=IFERROR(SUM(G3:G102)/SUM(E3:E102),0)")
    ws.cell(row=104, column=8).number_format = FMT_DECIMAL
    style_total_row(ws, 104, 9)

    set_col_widths(ws, {
        "A": 15, "B": 16, "C": 22, "D": 20, "E": 18,
        "F": 18, "G": 18, "H": 16, "I": 25
    })
    ws.freeze_panes = "A3"
    return ws


def build_dashboard(wb):
    """Sheet 1: Dashboard (لوحة التحكم) — must be built LAST to reference other sheets."""
    ws = wb.create_sheet("Dashboard", 0)  # Insert at position 0

    # Title
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = "📊 لوحة التحكم — 3D Makers Dashboard"
    title.font = Font(name="Arial", bold=True, color=WHITE, size=16)
    title.fill = FILL_HEADER
    title.alignment = ALIGN_CENTER

    # Subtitle
    ws.merge_cells("A2:D2")
    sub = ws["A2"]
    sub.value = "شراكة محمد (رأس المال) & نبيل (التشغيل والتسويق)"
    sub.font = Font(name="Arial", italic=True, color="78909C", size=10)
    sub.alignment = ALIGN_CENTER

    # Helper to add a KPI row
    def kpi_row(row, label, formula, fmt=FMT_CURRENCY, is_green=True):
        ws.cell(row=row, column=2, value=label).font = FONT_KPI_LABEL
        ws.cell(row=row, column=2).alignment = ALIGN_LEFT
        val_cell = ws.cell(row=row, column=3, value=formula)
        val_cell.font = FONT_KPI_VALUE_GREEN if is_green else FONT_KPI_VALUE
        val_cell.number_format = fmt
        val_cell.alignment = ALIGN_RIGHT
        val_cell.border = BOTTOM_BORDER

    def section_header(row, text):
        ws.merge_cells(f"A{row}:D{row}")
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = FONT_SECTION
        cell.fill = FILL_SECTION_BG
        cell.alignment = ALIGN_LEFT

    # ─── Section: Capital ───
    section_header(4, "💰 رأس المال / Capital")
    kpi_row(5, "إجمالي رأس المال المُستثمر", "=Capital!D103")
    kpi_row(6, "ثمن الطابعة ثلاثية الأبعاد", "=Capital!D3")

    # ─── Section: Revenue ───
    section_header(8, "🛒 الإيرادات والمبيعات / Revenue & Sales")
    kpi_row(9, "إجمالي الإيرادات حتى الآن", "=Sales!E203")
    kpi_row(10, "عدد القطع المباعة", "=Sales!C203", fmt=FMT_NUMBER, is_green=False)
    kpi_row(11, "متوسط سعر البيع لكل قطعة", "=IFERROR(Sales!E203/Sales!C203,0)")

    # ─── Section: Costs ───
    section_header(13, "📦 التكاليف / Costs")
    kpi_row(14, "إجمالي مصاريف المواد (المشتريات)", "=Purchases!D203")
    kpi_row(15, "إجمالي المصاريف العامة (التشغيلية)", "=Expenses!D203")
    kpi_row(16, "إجمالي مصاريف الإعلانات", "=Advertising!E103")
    kpi_row(17, "إجمالي التكاليف التشغيلية", "=Purchases!D203+Expenses!D203+Advertising!E103")

    # ─── Section: Production Cost Average ───
    # Average production cost per unit (from Product Calculator - average of filled rows)
    kpi_row(18, "متوسط تكلفة الإنتاج لكل قطعة",
            "=IFERROR(AVERAGEIF(ProductCalc!G7:G56,\">0\"),0)")

    # ─── Section: Profit ───
    section_header(20, "📈 الربح / Profit")
    kpi_row(21, "صافي الربح الإجمالي",
            "=Sales!E203-Purchases!D203-Expenses!D203-Advertising!E103")
    kpi_row(22, "نسبة هامش الربح الإجمالي",
            "=IFERROR((Sales!E203-Purchases!D203-Expenses!D203-Advertising!E103)/Sales!E203,0)",
            fmt=FMT_PERCENT, is_green=False)

    # ─── Section: Capital Recovery ───
    section_header(24, "🖨️ استرداد رأس المال / Capital Recovery")
    kpi_row(25, "الربح المتراكم (للاسترداد)",
            "=Sales!E203-Purchases!D203-Expenses!D203-Advertising!E103")
    kpi_row(26, "المبلغ المتبقي لاسترداد رأس المال",
            "=MAX(Capital!D103-(Sales!E203-Purchases!D203-Expenses!D203-Advertising!E103),0)")

    # Capital recovery status: Yes/No
    ws.cell(row=27, column=2, value="هل تم استرداد رأس المال؟").font = FONT_KPI_LABEL
    ws.cell(row=27, column=2).alignment = ALIGN_LEFT
    status_cell = ws.cell(row=27, column=3,
        value='=IF((Sales!E203-Purchases!D203-Expenses!D203-Advertising!E103)>=Capital!D103,"✅ نعم — تم الاسترداد","❌ لا — لم يتم بعد")')
    status_cell.font = Font(name="Arial", bold=True, color=DARK_HEADER_BG, size=12)
    status_cell.alignment = ALIGN_RIGHT
    status_cell.border = BOTTOM_BORDER

    kpi_row(28, "نسبة الاسترداد",
            "=IFERROR((Sales!E203-Purchases!D203-Expenses!D203-Advertising!E103)/Capital!D103,0)",
            fmt=FMT_PERCENT, is_green=False)

    # Column widths
    set_col_widths(ws, {"A": 3, "B": 40, "C": 25, "D": 5})

    # Background fill
    for r in range(1, 30):
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            if cell.fill == PatternFill():
                cell.fill = FILL_DASHBOARD_BG

    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = "1565C0"
    return ws


# ═══════════════════════════════════════════════════════════
# MAIN — Build the workbook
# ═══════════════════════════════════════════════════════════

def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Build sheets in order (Dashboard last because it references others, but inserted at pos 0)
    print("Building Capital sheet...")
    build_capital(wb)

    print("Building Product Cost Calculator sheet...")
    build_product_calculator(wb)

    print("Building Purchases sheet...")
    build_purchases(wb)

    print("Building Sales sheet...")
    build_sales(wb)

    print("Building Expenses sheet...")
    build_expenses(wb)

    print("Building Advertising sheet...")
    build_advertising(wb)

    print("Building Dashboard sheet...")
    build_dashboard(wb)

    # Set tab colors for visual distinction
    wb["Capital"].sheet_properties.tabColor = "FF8F00"
    wb["ProductCalc"].sheet_properties.tabColor = "7B1FA2"
    wb["Purchases"].sheet_properties.tabColor = "00838F"
    wb["Sales"].sheet_properties.tabColor = "2E7D32"
    wb["Expenses"].sheet_properties.tabColor = "D32F2F"
    wb["Advertising"].sheet_properties.tabColor = "F57C00"

    # Save
    wb.save(FILENAME)
    print(f"\n✅ Successfully created: {FILENAME}")
    print(f"   Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
